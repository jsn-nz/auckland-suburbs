#!/usr/bin/env python3
"""
Auckland suburb profiles - data pipeline.

Downloads (if not already present in data/raw/):
  1. Stats NZ 2023 Census totals by topic for individuals by SA2, parts 1 & 2
     (Datafinder layers 120897 / 120898) + their column lookup tables
     (documents 25543 / 25542) - CSV via WFS. Requires STATS_NZ_API_KEY.
  2. Geographic Areas Table 2023 (Datafinder table 111243) - meshblock-level
     concordance used to map SA2 -> region / local board / SA3.
  3. Statistical Area 2 2023 Clipped (generalised) boundaries
     (Datafinder layer 111206) as WGS84 GeoJSON.
  4. NZDep2023 SA2-level index (NZDep2023_WgtAvSA2.xlsx, University of Otago).
     Otago's WAF blocks non-browser clients; falls back to the Internet
     Archive's byte-identical capture of the same official file.

Then filters to Auckland region, joins everything on SA2 2023 code (never
name), computes derived metrics, and emits:
  docs/data/auckland.geojson - simplified geometry + core choropleth metrics
  docs/data/suburbs.json     - full per-SA2 detail + regional reference values

Data-correctness rules honoured here:
  * Ethnicity is multi-response: shares are computed against "total stated"
    and intentionally sum to >100%. Never normalised.
  * Counts are randomly rounded to base 3 by Stats NZ; not "fixed".
  * Suppressed cells arrive as -999 and become null; never estimated.
  * NZDep2023 decile 1 = LEAST deprived, 10 = most deprived.
  * If a download fails or the schema doesn't match expectations, the script
    aborts loudly. It never substitutes hardcoded figures.

Re-run:  python3 build.py           (uses cached raw files)
         python3 build.py --force   (re-downloads everything)
"""
import csv
import io
import json
import os
import subprocess
import sys
import urllib.request
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent
RAW = ROOT / "data" / "raw"
OUT = ROOT / "docs" / "data"
SUPPRESSED = -999  # Stats NZ confidentiality sentinel in these layers

csv.field_size_limit(sys.maxsize)

# ---------------------------------------------------------------- downloads

DATAFINDER = "https://datafinder.stats.govt.nz"

SOURCES = {
    "census_sa2_part1.csv": {
        "url": DATAFINDER + "/services;key={key}/wfs/layer-120897"
               "?service=WFS&version=2.0.0&request=GetFeature"
               "&typeNames=layer-120897&outputFormat=csv",
        "desc": "2023 Census totals by topic for individuals by SA2 - part 1 (layer 120897)",
    },
    "census_sa2_part2.csv": {
        "url": DATAFINDER + "/services;key={key}/wfs/layer-120898"
               "?service=WFS&version=2.0.0&request=GetFeature"
               "&typeNames=layer-120898&outputFormat=csv",
        "desc": "2023 Census totals by topic for individuals by SA2 - part 2 (layer 120898)",
    },
    "census_sa2_part1_lookup.csv": {
        "url": DATAFINDER + "/services/api/v1/documents/25543/versions/27388/download/?key={key}",
        "desc": "Part 1 column lookup table (document 25543)",
    },
    "census_sa2_part2_lookup.csv": {
        "url": DATAFINDER + "/services/api/v1/documents/25542/versions/27344/download/?key={key}",
        "desc": "Part 2 column lookup table (document 25542)",
    },
    "geographic_areas_2023.csv": {
        "url": DATAFINDER + "/services;key={key}/wfs/table-111243"
               "?service=WFS&version=2.0.0&request=GetFeature"
               "&typeNames=table-111243&outputFormat=csv",
        "desc": "Geographic Areas Table 2023 (table 111243)",
    },
    "census_sa2_dwellings.csv": {
        "url": DATAFINDER + "/services;key={key}/wfs/layer-120853"
               "?service=WFS&version=2.0.0&request=GetFeature"
               "&typeNames=layer-120853&outputFormat=csv",
        "desc": "2023 Census totals by topic for dwellings by SA2 (layer 120853)",
    },
    "census_sa2_dwellings_lookup.csv": {
        "url": DATAFINDER + "/services/api/v1/documents/25546/versions/{ver:25546}/download/?key={key}",
        "desc": "Dwellings column lookup table (document 25546)",
    },
    "census_sa2_households.csv": {
        "url": DATAFINDER + "/services;key={key}/wfs/layer-120892"
               "?service=WFS&version=2.0.0&request=GetFeature"
               "&typeNames=layer-120892&outputFormat=csv",
        "desc": "2023 Census totals by topic for households by SA2 (layer 120892)",
    },
    "census_sa2_households_lookup.csv": {
        "url": DATAFINDER + "/services/api/v1/documents/25548/versions/{ver:25548}/download/?key={key}",
        "desc": "Households column lookup table (document 25548)",
    },
    "sa2_2023_clipped_generalised.geojson": {
        "url": DATAFINDER + "/services;key={key}/wfs/layer-111206"
               "?service=WFS&version=2.0.0&request=GetFeature"
               "&typeNames=layer-111206&outputFormat=application/json"
               "&srsName=urn:ogc:def:crs:EPSG::4326",
        "desc": "Statistical Area 2 2023 Clipped (generalised) boundaries (layer 111206)",
    },
}

NZDEP_FILE = "NZDep2023_WgtAvSA2.xlsx"
NZDEP_LIVE = ("https://www.otago.ac.nz/__data/assets/excel_doc/0024/593142/"
              "NZDep2023_WgtAvSA2.xlsx")
NZDEP_ARCHIVE = ("https://web.archive.org/web/20241130120032id_/" + NZDEP_LIVE)


def die(msg):
    sys.exit("\nBUILD ABORTED: " + msg + "\nNothing was emitted. Fix the issue and re-run.")


def load_api_key():
    key = os.environ.get("STATS_NZ_API_KEY")
    if not key:
        env = ROOT / ".env"
        if env.exists():
            for line in env.read_text().splitlines():
                if line.startswith("STATS_NZ_API_KEY="):
                    key = line.split("=", 1)[1].strip()
    if not key:
        die("STATS_NZ_API_KEY not set (env var or .env file). Get a free key at "
            "https://datafinder.stats.govt.nz/my/api/")
    return key


def fetch(url, dest, min_bytes=1000):
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    try:
        with urllib.request.urlopen(req, timeout=600) as r:
            data = r.read()
    except Exception as e:
        raise RuntimeError("download failed: %s (%s)" % (url.split("?")[0], e))
    if len(data) < min_bytes or data[:5] in (b"<!DOC", b"<html"):
        raise RuntimeError("unexpected response (HTML error page or too small) from "
                           + url.split("?")[0])
    dest.write_bytes(data)
    return len(data)


def download_all(force=False):
    RAW.mkdir(parents=True, exist_ok=True)
    key = load_api_key()
    manifest = {}
    mpath = RAW / "manifest.json"
    if mpath.exists():
        manifest = json.loads(mpath.read_text())

    for name, src in SOURCES.items():
        dest = RAW / name
        if dest.exists() and not force:
            print("  cached   %s" % name)
            continue
        print("  fetching %s ..." % name)
        try:
            url = src["url"]
            if "{ver:" in url:  # resolve a document's latest version id
                doc_id = url.split("{ver:")[1].split("}")[0]
                vurl = ("%s/services/api/v1/documents/%s/versions/?key=%s"
                        % (DATAFINDER, doc_id, key))
                req = urllib.request.Request(vurl, headers={"User-Agent": "Mozilla/5.0"})
                with urllib.request.urlopen(req, timeout=60) as r:
                    ver = json.load(r)[0]["id"]
                url = url.replace("{ver:%s}" % doc_id, str(ver))
            n = fetch(url.format(key=key), dest)
        except RuntimeError as e:
            die(str(e))
        except Exception as e:
            die("could not resolve document version for %s (%s)" % (name, e))
        manifest[name] = {"downloaded": date.today().isoformat(),
                          "bytes": n, "source": src["desc"]}

    dest = RAW / NZDEP_FILE
    if not dest.exists() or force:
        print("  fetching %s ..." % NZDEP_FILE)
        used = NZDEP_LIVE
        try:
            n = fetch(NZDEP_LIVE, dest, min_bytes=50000)
        except RuntimeError:
            print("           otago.ac.nz blocked the request; using the Internet "
                  "Archive capture of the same file")
            used = NZDEP_ARCHIVE
            try:
                n = fetch(NZDEP_ARCHIVE, dest, min_bytes=50000)
            except RuntimeError as e:
                die(str(e))
        manifest[NZDEP_FILE] = {"downloaded": date.today().isoformat(),
                                "bytes": n,
                                "source": "NZDep2023 SA2 weighted averages, "
                                          "University of Otago (" + used + ")"}
    else:
        print("  cached   %s" % NZDEP_FILE)

    for name in list(SOURCES) + [NZDEP_FILE]:
        manifest.setdefault(name, {"downloaded": date.today().isoformat(),
                                   "bytes": (RAW / name).stat().st_size,
                                   "source": SOURCES.get(name, {}).get("desc", NZDEP_LIVE)})
    mpath.write_text(json.dumps(manifest, indent=2))
    return manifest

# ------------------------------------------------------------ column checks

# Column code -> (lookup Year, lookup Variable1, lookup Variable1_category).
# Verified against the published lookup tables at build time; the build stops
# if the lookup no longer agrees (i.e. Stats NZ re-issued the layer).
ETH_CATS = ["European", "Māori", "Pacific Peoples", "Asian",
            "Middle Eastern/Latin American/African", "Other Ethnicity", "Total stated"]
AGE_CATS = ["Under 15 years", "15-29 years", "30-64 years", "65 years and over", "Total"]

PART1_COLS = {
    "VAR_1_1":   ("2013", "Census usually resident population count", "Total"),
    "VAR_1_2":   ("2018", "Census usually resident population count", "Total"),
    "VAR_1_3":   ("2023", "Census usually resident population count", "Total"),
    # median age 2013/2018/2023
    "VAR_1_27":  ("2013", "Age", "Median"),
    "VAR_1_48":  ("2018", "Age", "Median"),
    "VAR_1_69":  ("2023", "Age", "Median"),
    # birthplace + languages, 2023
    "VAR_1_95":  ("2023", "Birthplace (NZ born/overseas born)", "NZ born"),
    "VAR_1_96":  ("2023", "Birthplace (NZ born/overseas born)", "Overseas born"),
    "VAR_1_99":  ("2023", "Birthplace (NZ born/overseas born)", "Total stated"),
    "VAR_1_206": ("2023", "Languages spoken (total responses)", "Māori"),
    "VAR_1_207": ("2023", "Languages spoken (total responses)", "Samoan"),
    "VAR_1_213": ("2023", "Languages spoken (total responses)", "Total stated"),
}
# age life-cycle bands: 2013 = VAR_1_70..74, 2018 = 75..79, 2023 = 80..84
for year, base in [("2013", 70), ("2018", 75), ("2023", 80)]:
    for ci, cat in enumerate(AGE_CATS):
        PART1_COLS["VAR_1_%d" % (base + ci)] = (year, "Age (life cycle groups)", cat)
# ethnicity: 2013 = VAR_1_136.., 2018 = 147.., 2023 = 158.. ("Total stated" is +10)
for year, base in [("2013", 136), ("2018", 147), ("2023", 158)]:
    for ci, cat in enumerate(ETH_CATS[:-1]):
        PART1_COLS["VAR_1_%d" % (base + ci)] = (year, "Ethnicity (total responses)", cat)
    PART1_COLS["VAR_1_%d" % (base + 10)] = (year, "Ethnicity (total responses)", "Total stated")

PART2_COLS = {
    # individual home ownership: 2013 base 1, 2018 base 7, 2023 base 13
    **{"VAR_2_%d" % (b + o): (y, "Individual home ownership", c)
       for y, b in [("2013", 1), ("2018", 7), ("2023", 13)]
       for o, c in [(0, "Hold in a family trust"), (1, "Own or partly own"),
                    (5, "Total stated")]},
    # highest qualification bachelor+ : 2013 base 208, 2018 base 223, 2023 base 238
    **{"VAR_2_%d" % (b + o): (y, "Highest qualification", c)
       for y, b in [("2013", 208), ("2018", 223), ("2023", 238)]
       for o, c in [(7, "Bachelor degree and Level 7 qualification"),
                    (8, "Post-graduate and honours degrees"),
                    (9, "Masters degree"), (10, "Doctorate degree"),
                    (14, "Total stated")]},
    # median personal income 2013/2018/2023
    "VAR_2_382": ("2013", "Total personal income", "Median ($)"),
    "VAR_2_393": ("2018", "Total personal income", "Median ($)"),
    "VAR_2_404": ("2023", "Total personal income", "Median ($)"),
    # usual residence 5 years ago, 2023
    "VAR_2_61":  ("2023", "Usual residence 5 years ago indicator", "Same as usual residence"),
    "VAR_2_64":  ("2023", "Usual residence 5 years ago indicator", "Overseas"),
    "VAR_2_71":  ("2023", "Usual residence 5 years ago indicator", "Total stated"),
    # work and labour force status, 2023
    "VAR_2_429": ("2023", "Work and labour force status", "Employed Full-time"),
    "VAR_2_430": ("2023", "Work and labour force status", "Employed Part-time"),
    "VAR_2_431": ("2023", "Work and labour force status", "Unemployed"),
    "VAR_2_432": ("2023", "Work and labour force status", "Not in the Labour Force"),
    "VAR_2_435": ("2023", "Work and labour force status", "Total stated"),
    # main means of travel to work (usual residence), 2023
    "VAR_2_773": ("2023", "Main means of travel to work, by usual residence address", "Work at home"),
    "VAR_2_774": ("2023", "Main means of travel to work, by usual residence address", "Did not go to work today"),
    "VAR_2_775": ("2023", "Main means of travel to work, by usual residence address", "Drive a private car, truck or van"),
    "VAR_2_776": ("2023", "Main means of travel to work, by usual residence address", "Drive a company car, truck or van"),
    "VAR_2_777": ("2023", "Main means of travel to work, by usual residence address", "Passenger in a car, truck, van or company bus"),
    "VAR_2_778": ("2023", "Main means of travel to work, by usual residence address", "Public bus"),
    "VAR_2_779": ("2023", "Main means of travel to work, by usual residence address", "Train"),
    "VAR_2_780": ("2023", "Main means of travel to work, by usual residence address", "Bicycle"),
    "VAR_2_781": ("2023", "Main means of travel to work, by usual residence address", "Walk or jog"),
    "VAR_2_782": ("2023", "Main means of travel to work, by usual residence address", "Ferry"),
    "VAR_2_783": ("2023", "Main means of travel to work, by usual residence address", "Other"),
    "VAR_2_786": ("2023", "Main means of travel to work, by usual residence address", "Total stated"),
}
# dwellings layer (VAR_3_*), 2023
PART3_COLS = {
    "VAR_3_29": ("2023", "Dwelling dampness", "Always damp"),
    "VAR_3_30": ("2023", "Dwelling dampness", "Sometimes damp"),
    "VAR_3_34": ("2023", "Dwelling dampness", "Total stated"),
    "VAR_3_41": ("2023", "Dwelling mould", "Mould over A4 size - always"),
    "VAR_3_42": ("2023", "Dwelling mould", "Mould over A4 size - sometimes"),
    "VAR_3_46": ("2023", "Dwelling mould", "Total stated"),
    "VAR_3_80": ("2023", "Dwelling type", "Private dwelling"),
    "VAR_3_82": ("2023", "Dwelling type", "Separate house"),
    "VAR_3_83": ("2023", "Dwelling type", "Joined dwelling"),
}
# households layer (VAR_4_*), 2023
PART4_COLS = {
    "VAR_4_48":  ("2023", "Household crowding index", "Crowded"),
    "VAR_4_57":  ("2023", "Household crowding index", "Total stated"),
    "VAR_4_225": ("2023", "Total household income", "Median ($)"),
    "VAR_4_261": ("2023", "Weekly rent paid by household", "Median ($)"),
    "VAR_4_260": ("2023", "Weekly rent paid by household", "Total stated"),
}


def verify_lookup(lookup_csv, expected):
    with open(lookup_csv, encoding="utf-8-sig") as f:
        rows = {r["Column_name"]: r for r in csv.DictReader(f)}
    for col, (year, var, cat) in expected.items():
        r = rows.get(col)
        if r is None or r["Year"] != year or r["Variable1"] != var \
                or r["Variable1_category"] != cat:
            got = ("%s | %s | %s" % (r["Year"], r["Variable1"],
                   r["Variable1_category"])) if r else "column missing"
            die("lookup table %s no longer matches for %s.\n  expected: %s | %s | %s"
                "\n  got:      %s\nStats NZ may have re-issued the layer; the column "
                "map in build.py needs re-verifying." % (
                    lookup_csv.name, col, year, var, cat, got))

# ---------------------------------------------------------------- parsing


def val(row, idx, col):
    """Parse one census cell -> float, or None when suppressed/blank."""
    v = row[idx[col]]
    if v in ("", "S", None):
        return None
    f = float(v)
    if f < 0:  # -999 confidential; treat any sentinel as suppressed
        return None
    return f


def read_census(path, cols):
    out = {}
    with open(path) as f:
        r = csv.reader(f)
        hdr = next(r)
        idx = {h: i for i, h in enumerate(hdr)}
        for need in ["SA22023_V1_00", "SA22023_V1_00_NAME"] + list(cols):
            if need not in idx:
                die("column %s missing from %s - schema changed" % (need, path.name))
        for row in r:
            code = row[idx["SA22023_V1_00"]]
            out[code] = {c: val(row, idx, c) for c in cols}
            out[code]["name"] = row[idx["SA22023_V1_00_NAME"]]
    return out


def read_geo_areas(path):
    """meshblock rows -> SA2 code -> {region, local board (modal), SA3 name}."""
    from collections import Counter, defaultdict
    boards = defaultdict(Counter)
    regions = defaultdict(Counter)
    sa3 = {}
    names = {}
    with open(path) as f:
        r = csv.DictReader(f)
        for need in ("SA22023_code", "REGC2023_name", "TALB2023_name", "SA32023_name"):
            if need not in r.fieldnames:
                die("column %s missing from geographic areas table - schema changed" % need)
        for row in r:
            code = row["SA22023_code"]
            regions[code][row["REGC2023_name"]] += 1
            boards[code][row["TALB2023_name"]] += 1
            sa3[code] = row["SA32023_name"]
            names[code] = row["SA22023_name"]
    return {code: {
        "region": regions[code].most_common(1)[0][0],
        "board": boards[code].most_common(1)[0][0].replace(" Local Board Area", ""),
        "sa3": sa3[code],
        "name": names[code],
    } for code in regions}


def read_nzdep(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    hdr = next(rows)
    expect = ("SA22023_code", "SA22023_name", "SA2_average_NZDep2023",
              "SA2_average_NZDep2023_score")
    if tuple(hdr[:4]) != expect:
        die("NZDep2023 xlsx header changed: got %s, expected %s" % (hdr[:4], expect))
    out = {}
    for r in rows:
        if r[0] is None:
            continue
        code = str(r[0])
        decile = int(r[2]) if r[2] is not None else None
        score = int(r[3]) if r[3] is not None else None
        if decile is not None and not 1 <= decile <= 10:
            die("NZDep decile out of range for SA2 %s: %s" % (code, decile))
        out[code] = {"decile": decile, "score": score}
    return out

# ---------------------------------------------------------------- derive


def pct(n, d):
    if n is None or d is None or d == 0:
        return None
    return round(100.0 * n / d, 1)


def weighted_median(pairs):
    """pairs of (value, weight); population-weighted median of SA2 medians."""
    pairs = sorted((p for p in pairs if p[0] is not None and p[1]), key=lambda p: p[0])
    total = sum(w for _, w in pairs)
    if not total:
        return None
    acc = 0
    for v, w in pairs:
        acc += w
        if acc >= total / 2.0:
            return v
    return pairs[-1][0]


def build():
    force = "--force" in sys.argv
    print("downloading raw data ->", RAW)
    manifest = download_all(force)

    print("verifying column lookup tables against the published definitions")
    verify_lookup(RAW / "census_sa2_part1_lookup.csv", PART1_COLS)
    verify_lookup(RAW / "census_sa2_part2_lookup.csv", PART2_COLS)
    verify_lookup(RAW / "census_sa2_dwellings_lookup.csv", PART3_COLS)
    verify_lookup(RAW / "census_sa2_households_lookup.csv", PART4_COLS)

    print("reading census individuals parts 1+2, dwellings, households")
    p1 = read_census(RAW / "census_sa2_part1.csv", PART1_COLS)
    p2 = read_census(RAW / "census_sa2_part2.csv", PART2_COLS)
    p3 = read_census(RAW / "census_sa2_dwellings.csv", PART3_COLS)
    p4 = read_census(RAW / "census_sa2_households.csv", PART4_COLS)
    print("reading geographic areas concordance")
    geo = read_geo_areas(RAW / "geographic_areas_2023.csv")
    print("reading NZDep2023")
    dep = read_nzdep(RAW / NZDEP_FILE)

    akl_codes = sorted(c for c, g in geo.items() if g["region"] == "Auckland Region")
    if not 500 <= len(akl_codes) <= 700:
        die("expected roughly 500-700 Auckland SA2s, got %d - concordance filter "
            "looks wrong" % len(akl_codes))
    print("Auckland region SA2s: %d" % len(akl_codes))

    missing = [c for c in akl_codes if c not in p1 or c not in p2]
    if missing:
        die("%d Auckland SA2s missing from census layers, e.g. %s" %
            (len(missing), missing[:5]))

    def sum_(*vals):
        return None if any(v is None for v in vals) else sum(vals)

    ETH_KEYS = ["European", "Māori", "Pacific", "Asian", "MELAA", "Other"]
    ETH_BASE = {"2013": 136, "2018": 147, "2023": 158}   # +10 = Total stated
    AGE_BASE = {"2013": 70, "2018": 75, "2023": 80}      # +4 = Total
    AGE_KEYS = ["0-14", "15-29", "30-64", "65+"]
    MEDAGE = {"2013": "VAR_1_27", "2018": "VAR_1_48", "2023": "VAR_1_69"}
    MEDINC = {"2013": "VAR_2_382", "2018": "VAR_2_393", "2023": "VAR_2_404"}
    OWN_BASE = {"2013": 1, "2018": 7, "2023": 13}        # trust, own, +5 stated
    QUAL_BASE = {"2013": 208, "2018": 223, "2023": 238}  # +7..+10 bach+, +14 stated
    POP = {"2013": "VAR_1_1", "2018": "VAR_1_2", "2023": "VAR_1_3"}

    def eth_year(a, year):
        b = ETH_BASE[year]
        stated = a["VAR_1_%d" % (b + 10)]
        return {k: {"n": a["VAR_1_%d" % (b + i)],
                    "pct": pct(a["VAR_1_%d" % (b + i)], stated)}
                for i, k in enumerate(ETH_KEYS)}, stated

    def age_year(a, year):
        b = AGE_BASE[year]
        tot = a["VAR_1_%d" % (b + 4)]
        return {k: {"n": a["VAR_1_%d" % (b + i)],
                    "pct": pct(a["VAR_1_%d" % (b + i)], tot)}
                for i, k in enumerate(AGE_KEYS)}

    def own_pct_year(b_, year):
        # "hold in a family trust" was not a 2013 category (-997 for every SA2)
        # and can be suppressed for tiny counts elsewhere - treat as 0 when
        # absent rather than voiding the whole rate
        base = OWN_BASE[year]
        own = b_["VAR_2_%d" % (base + 1)]
        trust = b_["VAR_2_%d" % base] or 0
        return pct(None if own is None else own + trust, b_["VAR_2_%d" % (base + 5)])

    def bach_pct_year(b_, year):
        base = QUAL_BASE[year]
        return pct(sum_(*[b_["VAR_2_%d" % (base + o)] for o in (7, 8, 9, 10)]),
                   b_["VAR_2_%d" % (base + 14)])

    # travel-to-work display groups -> part 2 column offsets (from VAR_2_)
    TRAVEL = [
        ("Work from home", ["VAR_2_773"]),
        ("Drive", ["VAR_2_775", "VAR_2_776"]),
        ("Passenger", ["VAR_2_777"]),
        ("Public transport", ["VAR_2_778", "VAR_2_779", "VAR_2_782"]),  # bus+train+ferry
        ("Cycle", ["VAR_2_780"]),
        ("Walk or jog", ["VAR_2_781"]),
        ("Other", ["VAR_2_783"]),
        # "Did not go to work today" (VAR_2_774) is -997 for every SA2:
        # it was a 2018 category that no longer exists in the 2023 question
    ]

    suburbs = []
    for code in akl_codes:
        a, b, g = p1[code], p2[code], geo[code]
        c3 = p3.get(code, {})
        c4 = p4.get(code, {})
        d = dep.get(code, {"decile": None, "score": None})
        pop23, pop18 = a["VAR_1_3"], a["VAR_1_2"]
        eth2023, eth_stated = eth_year(a, "2023")
        age_total = a["VAR_1_84"]
        own_num = None if b["VAR_2_14"] is None else b["VAR_2_14"] + (b["VAR_2_13"] or 0)
        bach_num = sum_(b["VAR_2_245"], b["VAR_2_246"], b["VAR_2_247"], b["VAR_2_248"])
        travel_stated = b["VAR_2_786"]
        lf = sum_(b["VAR_2_429"], b["VAR_2_430"], b["VAR_2_431"])  # labour force
        sa3 = g["sa3"]

        hist = {}
        for year in ("2013", "2018"):
            eth_y, _ = eth_year(a, year)
            hist[year] = {
                "pop": None if a[POP[year]] is None else int(a[POP[year]]),
                "median_age": a[MEDAGE[year]],
                "median_income": b[MEDINC[year]],
                "home_own_pct": own_pct_year(b, year),
                "bachelor_pct": bach_pct_year(b, year),
                "ethnicity": {k: v["pct"] for k, v in eth_y.items()},
                "age": {k: v["pct"] for k, v in age_year(a, year).items()},
            }
        s = {
            "code": code,
            "name": g["name"],
            "board": g["board"],
            # colloquial grouping: official SA3 2023 name, only where it
            # differs from the SA2 name (i.e. this SA2 is part of a broader
            # commonly-named area). Otherwise blank - never guessed.
            "colloquial": sa3 if sa3 != g["name"] else "",
            "pop2023": None if pop23 is None else int(pop23),
            "pop2018": None if pop18 is None else int(pop18),
            "pop_change_pct": pct((pop23 - pop18) if (pop23 is not None and pop18 is not None) else None, pop18),
            "median_age": a["VAR_1_69"],
            "age": {  # counts + shares of the age-group total
                "0-14":  {"n": a["VAR_1_80"], "pct": pct(a["VAR_1_80"], age_total)},
                "15-29": {"n": a["VAR_1_81"], "pct": pct(a["VAR_1_81"], age_total)},
                "30-64": {"n": a["VAR_1_82"], "pct": pct(a["VAR_1_82"], age_total)},
                "65+":   {"n": a["VAR_1_83"], "pct": pct(a["VAR_1_83"], age_total)},
            },
            # multi-response: shares of people who stated an ethnicity;
            # groups overlap so these intentionally sum to >100%
            "ethnicity": eth2023,
            "eth_stated": None if eth_stated is None else int(eth_stated),
            "median_income": b["VAR_2_404"],
            "bachelor_pct": pct(bach_num, b["VAR_2_252"]),
            "home_own_pct": pct(own_num, b["VAR_2_18"]),
            "dep_decile": d["decile"],   # 1 = least deprived, 10 = most
            "dep_score": d["score"],
            "pop2013": None if a["VAR_1_1"] is None else int(a["VAR_1_1"]),
            "hist": hist,
            # travel to work: shares of everyone who stated a travel mode
            # (includes people who worked from home or didn't go that day)
            "travel": {label: {"n": sum_(*[b[c] for c in cols]),
                               "pct": pct(sum_(*[b[c] for c in cols]), travel_stated)}
                       for label, cols in TRAVEL},
            "unemployment_pct": pct(b["VAR_2_431"], lf),
            "overseas_born_pct": pct(a["VAR_1_96"], a["VAR_1_99"]),
            "te_reo_pct": pct(a["VAR_1_206"], a["VAR_1_213"]),
            "samoan_pct": pct(a["VAR_1_207"], a["VAR_1_213"]),
            "same_home_5y_pct": pct(b["VAR_2_61"], b["VAR_2_71"]),
            "damp_pct": pct(sum_(c3.get("VAR_3_29"), c3.get("VAR_3_30")),
                            c3.get("VAR_3_34")),
            "mould_pct": pct(sum_(c3.get("VAR_3_41"), c3.get("VAR_3_42")),
                             c3.get("VAR_3_46")),
            "separate_house_pct": pct(c3.get("VAR_3_82"), c3.get("VAR_3_80")),
            "crowded_pct": pct(c4.get("VAR_4_48"), c4.get("VAR_4_57")),
            "median_hh_income": c4.get("VAR_4_225"),
            "median_rent": c4.get("VAR_4_261"),
        }
        suburbs.append(s)

    n_dep = sum(1 for s in suburbs if s["dep_decile"] is not None)
    print("SA2s with an NZDep2023 decile: %d / %d (rest are non-residential "
          "areas such as inlets/islands/ports)" % (n_dep, len(suburbs)))

    # ---- regional reference values (aggregated from the Auckland SA2 rows)
    def total(key_fn):
        vals = [key_fn(s) for s in suburbs]
        return sum(v for v in vals if v is not None)

    reg_pop23 = total(lambda s: s["pop2023"])
    reg_pop18 = total(lambda s: s["pop2018"])
    reg_eth_stated = total(lambda s: s["eth_stated"])
    reg_age_total = total(lambda s: sum(a["n"] for a in s["age"].values())
                          if all(a["n"] is not None for a in s["age"].values()) else None)
    region = {
        "pop2023": reg_pop23,
        "pop_change_pct": pct(reg_pop23 - reg_pop18, reg_pop18),
        # population-weighted median of SA2 medians (no region-level medians
        # are published in this dataset) - labelled as derived in the UI
        "median_age": weighted_median([(s["median_age"], s["pop2023"]) for s in suburbs]),
        "median_income": weighted_median([(s["median_income"], s["pop2023"]) for s in suburbs]),
        "ethnicity": {k: pct(total(lambda s, k=k: s["ethnicity"][k]["n"]), reg_eth_stated)
                      for k in ("European", "Māori", "Pacific", "Asian", "MELAA", "Other")},
        "age": {k: pct(total(lambda s, k=k: s["age"][k]["n"]), reg_age_total)
                for k in ("0-14", "15-29", "30-64", "65+")},
        "bachelor_pct": pct(
            sum(p2[s["code"]]["VAR_2_245"] + p2[s["code"]]["VAR_2_246"] +
                p2[s["code"]]["VAR_2_247"] + p2[s["code"]]["VAR_2_248"]
                for s in suburbs if s["bachelor_pct"] is not None),
            sum(p2[s["code"]]["VAR_2_252"] for s in suburbs
                if s["bachelor_pct"] is not None)),
        "home_own_pct": pct(
            sum((p2[s["code"]]["VAR_2_13"] or 0) + p2[s["code"]]["VAR_2_14"]
                for s in suburbs if s["home_own_pct"] is not None),
            sum(p2[s["code"]]["VAR_2_18"] for s in suburbs
                if s["home_own_pct"] is not None)),
    }

    # regional shares for the new topics: sum numerators/denominators over the
    # SA2s where the suburb-level share could be computed
    def agg_pct(metric_key, num_fn, den_fn):
        num = den = 0
        for s in suburbs:
            if s[metric_key] is None:
                continue
            num += num_fn(s["code"])
            den += den_fn(s["code"])
        return pct(num, den)

    # per category: aggregate over the SA2s where that category + the
    # denominator are both published (suppression differs per category)
    region["travel"] = {}
    for label, cols in TRAVEL:
        num = den = 0
        for s in suburbs:
            vals = [p2[s["code"]][c] for c in cols]
            stated = p2[s["code"]]["VAR_2_786"]
            if stated is None or any(v is None for v in vals):
                continue
            num += sum(vals)
            den += stated
        region["travel"][label] = pct(num, den)
    region["unemployment_pct"] = agg_pct(
        "unemployment_pct", lambda c: p2[c]["VAR_2_431"],
        lambda c: p2[c]["VAR_2_429"] + p2[c]["VAR_2_430"] + p2[c]["VAR_2_431"])
    region["overseas_born_pct"] = agg_pct(
        "overseas_born_pct", lambda c: p1[c]["VAR_1_96"], lambda c: p1[c]["VAR_1_99"])
    region["te_reo_pct"] = agg_pct(
        "te_reo_pct", lambda c: p1[c]["VAR_1_206"], lambda c: p1[c]["VAR_1_213"])
    region["samoan_pct"] = agg_pct(
        "samoan_pct", lambda c: p1[c]["VAR_1_207"], lambda c: p1[c]["VAR_1_213"])
    region["same_home_5y_pct"] = agg_pct(
        "same_home_5y_pct", lambda c: p2[c]["VAR_2_61"], lambda c: p2[c]["VAR_2_71"])
    region["damp_pct"] = agg_pct(
        "damp_pct", lambda c: p3[c]["VAR_3_29"] + p3[c]["VAR_3_30"],
        lambda c: p3[c]["VAR_3_34"])
    region["mould_pct"] = agg_pct(
        "mould_pct", lambda c: p3[c]["VAR_3_41"] + p3[c]["VAR_3_42"],
        lambda c: p3[c]["VAR_3_46"])
    region["separate_house_pct"] = agg_pct(
        "separate_house_pct", lambda c: p3[c]["VAR_3_82"], lambda c: p3[c]["VAR_3_80"])
    region["crowded_pct"] = agg_pct(
        "crowded_pct", lambda c: p4[c]["VAR_4_48"], lambda c: p4[c]["VAR_4_57"])
    # weighted medians of SA2 medians (weight = households with a stated value,
    # approximated by the crowding/rent "total stated" counts)
    region["median_hh_income"] = weighted_median(
        [(s["median_hh_income"], p4.get(s["code"], {}).get("VAR_4_57")) for s in suburbs])
    region["median_rent"] = weighted_median(
        [(s["median_rent"], p4.get(s["code"], {}).get("VAR_4_260")) for s in suburbs])
    # historical regional references for the time toggle
    region["hist"] = {}
    for year in ("2013", "2018"):
        eb, ab = ETH_BASE[year], AGE_BASE[year]
        eth_stated_y = sum(p1[s["code"]]["VAR_1_%d" % (eb + 10)] or 0 for s in suburbs)
        age_tot_y = sum(p1[s["code"]]["VAR_1_%d" % (ab + 4)] or 0 for s in suburbs)
        region["hist"][year] = {
            "median_age": weighted_median(
                [(s["hist"][year]["median_age"], s["hist"][year]["pop"]) for s in suburbs]),
            "median_income": weighted_median(
                [(s["hist"][year]["median_income"], s["hist"][year]["pop"]) for s in suburbs]),
            "home_own_pct": pct(
                sum((p2[s["code"]]["VAR_2_%d" % OWN_BASE[year]] or 0) +
                    (p2[s["code"]]["VAR_2_%d" % (OWN_BASE[year] + 1)] or 0) for s in suburbs),
                sum(p2[s["code"]]["VAR_2_%d" % (OWN_BASE[year] + 5)] or 0 for s in suburbs)),
            "bachelor_pct": pct(
                sum(sum(p2[s["code"]]["VAR_2_%d" % (QUAL_BASE[year] + o)] or 0
                        for o in (7, 8, 9, 10)) for s in suburbs),
                sum(p2[s["code"]]["VAR_2_%d" % (QUAL_BASE[year] + 14)] or 0 for s in suburbs)),
            "ethnicity": {k: pct(
                sum(p1[s["code"]]["VAR_1_%d" % (eb + i)] or 0 for s in suburbs), eth_stated_y)
                for i, k in enumerate(ETH_KEYS)},
            "age": {k: pct(
                sum(p1[s["code"]]["VAR_1_%d" % (ab + i)] or 0 for s in suburbs), age_tot_y)
                for i, k in enumerate(AGE_KEYS)},
        }

    OUT.mkdir(parents=True, exist_ok=True)

    # ---- geometry: filter to Auckland, attach core metrics, simplify
    print("filtering boundaries + attaching choropleth metrics")
    fc = json.loads((RAW / "sa2_2023_clipped_generalised.geojson").read_text())
    by_code = {s["code"]: s for s in suburbs}
    feats = []
    for f in fc["features"]:
        code = f["properties"].get("SA22023_V1_00")
        s = by_code.get(code)
        if s is None:
            continue
        f["properties"] = {
            "code": code, "name": s["name"], "board": s["board"],
            "dep_decile": s["dep_decile"], "median_age": s["median_age"],
            "median_income": s["median_income"], "pop2023": s["pop2023"],
            "pop_change_pct": s["pop_change_pct"], "home_own_pct": s["home_own_pct"],
            **{"eth_" + k: v["pct"] for k, v in s["ethnicity"].items()},
            "travel_pt": s["travel"]["Public transport"]["pct"],
            "travel_wfh": s["travel"]["Work from home"]["pct"],
            "travel_active": (None if (s["travel"]["Cycle"]["pct"] is None or
                                       s["travel"]["Walk or jog"]["pct"] is None)
                              else round(s["travel"]["Cycle"]["pct"] +
                                         s["travel"]["Walk or jog"]["pct"], 1)),
            "unemployment_pct": s["unemployment_pct"],
            "overseas_born_pct": s["overseas_born_pct"],
            "te_reo_pct": s["te_reo_pct"],
            "same_home_5y_pct": s["same_home_5y_pct"],
            "damp_pct": s["damp_pct"],
            "mould_pct": s["mould_pct"],
            "separate_house_pct": s["separate_house_pct"],
            "crowded_pct": s["crowded_pct"],
            "median_hh_income": s["median_hh_income"],
            "median_rent": s["median_rent"],
        }
        feats.append(f)
    have = {f["properties"]["code"] for f in feats}
    no_geom = [s for s in suburbs if s["code"] not in have]
    # The clipped boundary layer intentionally drops SA2s that are entirely
    # water (inlets/oceanic/bays). Those stay in suburbs.json but can't be
    # drawn. Anything populated going missing means the wrong layer.
    bad = [s for s in no_geom if (s["pop2023"] or 0) > 100]
    if bad:
        die("boundary file is missing populated Auckland SA2s: %s - wrong "
            "boundary layer?" % [(s["code"], s["name"]) for s in bad])
    for s in suburbs:
        s["on_map"] = s["code"] in have
    if no_geom:
        print("  %d water-only SA2s have no clipped geometry (kept in table, "
              "not on map): %s" % (len(no_geom),
              ", ".join(s["name"] for s in no_geom)))

    (OUT / "suburbs.json").write_text(json.dumps({
        "generated": date.today().isoformat(),
        "sources": manifest,
        "region": region,
        "suburbs": suburbs,
    }, ensure_ascii=False, separators=(",", ":")))
    print("wrote docs/data/suburbs.json (%d suburbs)" % len(suburbs))
    pre = OUT / "auckland_full.geojson"
    pre.write_text(json.dumps({"type": "FeatureCollection", "features": feats},
                              ensure_ascii=False, separators=(",", ":")))
    size_before = pre.stat().st_size

    print("simplifying with mapshaper (15%%, shapes preserved)")
    final = OUT / "auckland.geojson"
    try:
        subprocess.run(
            ["mapshaper", str(pre), "-simplify", "15%", "keep-shapes",
             "-o", "precision=0.00001", "format=geojson", "force", str(final)],
            check=True, capture_output=True, text=True)
    except FileNotFoundError:
        die("mapshaper not found - install with: npm install -g mapshaper")
    except subprocess.CalledProcessError as e:
        die("mapshaper failed:\n" + e.stderr)
    size_after = final.stat().st_size
    pre.unlink()
    print("geometry size: %.1f MB before -> %.1f MB after simplification"
          % (size_before / 1e6, size_after / 1e6))
    total_out = size_after + (OUT / "suburbs.json").stat().st_size
    print("total payload (auckland.geojson + suburbs.json): %.2f MB" % (total_out / 1e6))
    if size_after > 3_000_000:
        print("WARNING: geometry still over 3 MB - consider a lower simplify %")
    print("done.")


if __name__ == "__main__":
    build()
